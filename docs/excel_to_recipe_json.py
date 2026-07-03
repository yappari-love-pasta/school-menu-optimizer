#!/usr/bin/env python3
"""
学校給食献立表Excelをreciept.json形式に変換するスクリプト
対応フォーマット: 堺市小学校 形式の給食献立表 (2025年度〜)

使用方法:
    python excel_to_recipe_json.py <Excelファイルパス> [<Excelファイルパス2> ...] [-o 出力JSONパス]

例:
    python excel_to_recipe_json.py 4月献立表.xlsx
    python excel_to_recipe_json.py 4月*.xlsx 5月*.xlsx -o output.json

出力形式 (reciept.json のスキーマ):
    - id          : 通し番号
    - title       : 料理名
    - category    : 1=主食, 2=主菜, 3=副菜, 4=汁物, 5=デザート・飲み物
    - genre       : 0固定
    - steps       : 調理工程数（デフォルト5、Excelに情報なし）
    - nutritions  : {エネルギー, たんぱく質, 脂質, ナトリウム}
                    ※日合計を料理数で按分した概算値（ナトリウム・脂質は0）
    - ingredients : [{id, name, amount, energy, protein, lipid, sodium}]
                    ※その日の全食材を主菜に割り当て（energy等は0、Excelに情報なし）

注意:
    - 重複するレシピ名は初出のデータを使用します
    - 食材の詳細栄養素(energy/protein/lipid/sodium)はExcelにないため0になります
    - ナトリウムはExcelに記載がないため0になります（塩分相当量に変換）
"""

import argparse
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("エラー: openpyxlが必要です。pip install openpyxl でインストールしてください。")
    sys.exit(1)


# -----------------------------------------------------------------------
# カテゴリ定義
# -----------------------------------------------------------------------
CATEGORY_KEYWORDS = {
    1: ["ごはん", "米", "パン", "コッペ", "食パン", "ロールパン", "ナン",
        "めん", "麺", "スパゲティ", "ライス", "うどん", "そば",
        "ちゃんぽん", "ラーメン", "焼きそば", "チャーハン"],
    4: ["スープ", "みそ汁", "味噌汁", "汁", "ポタージュ", "シチュー",
        "おでん", "ブイヨン", "コンソメ汁"],
    5: ["牛乳", "ヨーグルト", "デザート", "ゼリー", "プリン",
        "アイス", "ジュース", "フルーツ"],
    3: ["サラダ", "おひたし", "和え", "なます", "ひじき", "きんぴら",
        "漬け", "酢", "炒め煮", "ごぼう"],
}


def infer_category(title: str) -> int:
    """料理名からカテゴリIDを推定（優先順位: 主食→汁物→デザート→副菜→主菜）"""
    for cat_id in [1, 4, 5, 3]:
        for kw in CATEGORY_KEYWORDS[cat_id]:
            if kw in title:
                return cat_id
    return 2  # デフォルト: 主菜


def clean_title(val) -> str:
    """料理名から先頭の記号・空白を除去"""
    s = str(val).strip()
    s = re.sub(r'^[☆★＊*◆◇●○■□▲△▼▽・\s]+', '', s)
    return s.strip()


# -----------------------------------------------------------------------
# セルアクセスユーティリティ
# -----------------------------------------------------------------------
def cell_value(ws, row: int, col: int):
    """マージセルを考慮してセル値を返す（Noneの場合はNone）"""
    c = ws.cell(row=row, column=col)
    if c.value is not None:
        return c.value
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            return ws.cell(row=mr.min_row, column=mr.min_col).value
    return None


def cell_str(ws, row: int, col: int) -> str:
    v = cell_value(ws, row, col)
    return str(v).strip() if v is not None else ""


def is_date_num(val) -> bool:
    try:
        n = int(str(val).strip())
        return 1 <= n <= 31
    except (ValueError, TypeError):
        return False


def try_float(val) -> float | None:
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


# -----------------------------------------------------------------------
# Excel解析
# -----------------------------------------------------------------------
EXCLUDE_RECIPE_WORDS = {
    "こんだて", "献立", "栄養素", "栄養", "おかず", "内容",
    "エネルギー", "たんぱく質", "脂質", "脂肪", "ナトリウム",
    "塩分", "カルシウム", "鉄", "ビタミン",
    "主食", "主菜", "副菜", "汁物", "デザート",
    "赤", "緑", "黄", "食品名", "食材",
}


def looks_like_recipe(val) -> bool:
    """セル値がレシピ名らしいか判定"""
    if val is None:
        return False
    s = str(val).strip()
    if len(s) < 2:
        return False
    # 数値のみ → NO
    if re.fullmatch(r'[\d.,\s]+', s):
        return False
    # 除外ワードと完全一致 → NO
    cleaned = clean_title(s)
    if cleaned in EXCLUDE_RECIPE_WORDS:
        return False
    # 除外ワードで始まる → NO（"栄養素..." など）
    for w in EXCLUDE_RECIPE_WORDS:
        if s.startswith(w):
            return False
    # 日本語が含まれる → YES
    if re.search(r'[ぁ-んァ-ン一-龯]', cleaned):
        return True
    return False


def parse_color_ingredient(val) -> tuple[str | None, str]:
    """
    '（赤）鶏肉' や '(緑)にんじん' のような文字列を (color_label, name) に分解
    color_labelがなければ (None, val)
    """
    if val is None:
        return None, ""
    s = str(val).strip()
    m = re.match(r'^[（(]([赤緑黄])[）)]\s*(.+)$', s)
    if m:
        return m.group(1), m.group(2).strip()
    return None, s


def find_kondaite_rows(ws) -> list[int]:
    """
    'こんだて' が含まれる行番号リストを返す（週ブロックの開始行）
    左端3列以内にある場合を対象とする
    """
    rows = []
    for row in range(1, ws.max_row + 1):
        for col in range(1, 4):
            v = cell_str(ws, row, col)
            if "こんだて" in v or "献立" in v:
                rows.append(row)
                break
    return rows


def find_date_columns(ws, start_row: int, end_row: int) -> dict[int, int]:
    """
    週ブロック内で日付数字(1〜31)が入っている列を探す
    Returns: {column_index: date_number}
    """
    result = {}
    for row in range(start_row, min(start_row + 5, end_row + 1)):
        for col in range(2, ws.max_column + 1):
            v = cell_value(ws, row, col)
            if is_date_num(v):
                result[col] = int(str(v).strip())
    return result


def find_nutrition_row(ws, start_row: int, end_row: int) -> int | None:
    """
    週ブロック内で栄養素データ行を探す
    '栄養' というラベルがある行 or エネルギー(300〜800)が複数現れる行
    """
    for row in range(start_row, min(start_row + 20, end_row + 1)):
        # 左端3列に「栄養」があるか
        for col in range(1, 4):
            v = cell_str(ws, row, col)
            if "栄養" in v:
                return row
        # 列に大きな数値(エネルギー相当)が複数あるか
        energy_count = 0
        for col in range(2, ws.max_column + 1):
            f = try_float(cell_value(ws, row, col))
            if f is not None and 200 <= f <= 1000:
                energy_count += 1
        if energy_count >= 3:  # 3日分以上エネルギー値があれば栄養素行とみなす
            return row
    return None


def extract_recipes_in_day(ws, row_start: int, row_end: int,
                             col_start: int, col_end: int) -> list[str]:
    """
    指定セル範囲内からレシピ名を抽出する
    """
    recipes = []
    seen = set()
    for row in range(row_start, row_end + 1):
        for col in range(col_start, col_end + 1):
            v = cell_value(ws, row, col)
            if looks_like_recipe(v):
                title = clean_title(str(v))
                if title and title not in seen:
                    seen.add(title)
                    recipes.append(title)
    return recipes


def extract_nutrition_for_day(ws, nutrition_row: int,
                               col_start: int, col_end: int) -> dict:
    """
    栄養素行から1日分の栄養データを抽出する
    同行に エネルギー > たんぱく質 > 脂肪エネルギー比 の順で並んでいることが多い
    """
    nutrition = {"エネルギー": 0.0, "たんぱく質": 0.0, "脂質": 0.0, "ナトリウム": 0.0}

    nums = []
    for col in range(col_start, col_end + 1):
        f = try_float(cell_value(ws, nutrition_row, col))
        if f is not None and f > 0:
            nums.append(f)

    if not nums:
        return nutrition

    # エネルギー(最大値), たんぱく質(中間値), 脂肪エネルギー比(小数値) を推定
    nums_sorted = sorted(nums, reverse=True)
    if nums_sorted[0] > 100:
        nutrition["エネルギー"] = nums_sorted[0]
    if len(nums_sorted) > 1 and nums_sorted[1] < 100:
        nutrition["たんぱく質"] = nums_sorted[1]
    # 脂質はExcelの脂肪エネルギー比(%)からおおよそ換算
    # 脂肪エネルギー比(%) = 脂質(g)*9 / エネルギー(kcal) * 100
    if len(nums_sorted) > 2 and nutrition["エネルギー"] > 0:
        fat_ratio = nums_sorted[2]
        if 5 <= fat_ratio <= 60:  # 脂肪エネルギー比として妥当な範囲
            nutrition["脂質"] = round(nutrition["エネルギー"] * fat_ratio / 100 / 9, 1)

    return nutrition


def extract_ingredients_for_day(ws, row_start: int, row_end: int,
                                  col_start: int, col_end: int) -> list[dict]:
    """
    食材セクションから食材名と量を抽出する
    フォーマット例: "(赤)鶏肉" / 55 / "(緑)にんじん" / 10 のように
    名前セルと量セルが交互に並ぶ場合が多い
    """
    ingredients = []
    seen_names = set()
    ing_id = 1

    for row in range(row_start, row_end + 1):
        col = col_start
        while col <= col_end:
            v = cell_value(ws, row, col)
            if v is None:
                col += 1
                continue

            color_label, name = parse_color_ingredient(v)

            # 名前セルらしい場合
            if name and re.search(r'[ぁ-んァ-ン一-龯a-zA-Zａ-ｚＡ-Ｚ]', name):
                # 除外ワードチェック
                if name.strip() in EXCLUDE_RECIPE_WORDS:
                    col += 1
                    continue
                if len(name.strip()) < 1:
                    col += 1
                    continue

                # 次の列に量があるか確認
                amount = 0.0
                if col + 1 <= col_end:
                    next_v = cell_value(ws, row, col + 1)
                    f = try_float(next_v)
                    if f is not None and 0 < f < 2000:
                        amount = f
                        col += 1  # 量の列を消費

                clean_name = name.strip()
                # ゼロ量のものは記録しない（食材名だけで量が空のセルは除外）
                if amount > 0 and clean_name not in seen_names:
                    seen_names.add(clean_name)
                    ingredients.append({
                        "id": ing_id,
                        "name": clean_name,
                        "lipid": 0.0,
                        "amount": amount,
                        "energy": 0,
                        "sodium": 0,
                        "protein": 0.0,
                    })
                    ing_id += 1

            col += 1

    return ingredients


def day_col_range(date_cols: dict[int, int], this_col: int,
                   max_col: int) -> tuple[int, int]:
    """
    日付列から、その日が占める列範囲(start, end)を推定する
    次の日付列-1 を終端とする
    """
    sorted_cols = sorted(date_cols.keys())
    idx = sorted_cols.index(this_col)
    col_start = this_col
    if idx + 1 < len(sorted_cols):
        col_end = sorted_cols[idx + 1] - 1
    else:
        col_end = max_col
    return col_start, col_end


# -----------------------------------------------------------------------
# メイン解析ロジック
# -----------------------------------------------------------------------
def parse_excel_file(filepath: str) -> list[dict]:
    """
    1つのExcelファイルを解析してレシピリスト(dict)を返す
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # 最初のシート(または「献立」という名前のシート)を対象とする
    ws = None
    for sheet_name in wb.sheetnames:
        if "献立" in sheet_name or "メニュー" in sheet_name:
            ws = wb[sheet_name]
            break
    if ws is None:
        ws = wb.active

    max_row = ws.max_row
    max_col = ws.max_column

    # ---- Step 1: こんだて行(週ブロックの先頭)を探す ----
    kondaite_rows = find_kondaite_rows(ws)

    if not kondaite_rows:
        print(f"  警告: 'こんだて'が見つかりません → {filepath}")
        print("  ヒント: Excelシートに 'こんだて' または '献立' というテキストが含まれるか確認してください")
        return []

    recipes_map: dict[str, dict] = {}  # title → recipe dict (重複排除)
    recipe_id_counter = [1]

    for week_idx, week_start in enumerate(kondaite_rows):
        # 次の週ブロック開始行(またはファイル末尾)
        if week_idx + 1 < len(kondaite_rows):
            week_end = kondaite_rows[week_idx + 1] - 1
        else:
            week_end = max_row

        # ---- Step 2: 日付列を特定 ----
        date_cols = find_date_columns(ws, week_start, week_end)
        if not date_cols:
            continue

        # ---- Step 3: 栄養素行を探す ----
        nutrition_row = find_nutrition_row(ws, week_start, week_end)

        # 料理名を探す行範囲
        recipe_row_end = (nutrition_row - 1) if nutrition_row else (week_start + 8)
        # 食材を探す行範囲
        ingr_row_start = (nutrition_row + 1) if nutrition_row else (week_start + 9)
        ingr_row_end = week_end

        # ---- Step 4: 各日を処理 ----
        for date_col, date_num in sorted(date_cols.items()):
            col_start, col_end = day_col_range(date_cols, date_col, max_col)

            # 料理名抽出
            recipes = extract_recipes_in_day(
                ws, week_start, recipe_row_end, col_start, col_end
            )

            # 栄養素抽出
            day_nutrition = {"エネルギー": 0.0, "たんぱく質": 0.0, "脂質": 0.0, "ナトリウム": 0.0}
            if nutrition_row:
                day_nutrition = extract_nutrition_for_day(ws, nutrition_row, col_start, col_end)

            # 食材抽出
            day_ingredients = extract_ingredients_for_day(
                ws, ingr_row_start, ingr_row_end, col_start, col_end
            )

            # 料理ごとのエントリを作成
            num_recipes = max(len(recipes), 1)
            for r_idx, title in enumerate(recipes):
                if not title:
                    continue
                if title in recipes_map:
                    continue  # 同名レシピは初出を使用

                # 栄養素を料理数で按分
                per_recipe_nutrition = {
                    "エネルギー": round(day_nutrition["エネルギー"] / num_recipes, 1),
                    "たんぱく質": round(day_nutrition["たんぱく質"] / num_recipes, 1),
                    "脂質": round(day_nutrition["脂質"] / num_recipes, 1),
                    "ナトリウム": 0.0,  # Excelに記載なし
                }

                # 食材は最初の料理(主菜扱い)に全て割り当てる
                ingredients = []
                if r_idx == 0 and day_ingredients:
                    ingredients = [
                        {**ing, "id": i + 1}
                        for i, ing in enumerate(day_ingredients)
                    ]

                recipes_map[title] = {
                    "id": recipe_id_counter[0],
                    "title": title,
                    "category": infer_category(title),
                    "genre": 0,
                    "steps": 5,
                    "nutritions": per_recipe_nutrition,
                    "ingredients": ingredients,
                }
                recipe_id_counter[0] += 1

    return list(recipes_map.values())


# -----------------------------------------------------------------------
# エントリーポイント
# -----------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="学校給食献立表ExcelをJSON(reciept.json形式)に変換します",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "excel_files",
        nargs="+",
        metavar="EXCEL",
        help="変換するExcelファイル (.xlsx/.xls)",
    )
    parser.add_argument(
        "-o", "--output",
        metavar="OUTPUT_JSON",
        help="出力JSONファイルパス (省略時: 入力ファイル名.json)",
    )
    args = parser.parse_args()

    all_recipes: list[dict] = []
    all_titles: set[str] = set()

    for excel_path_str in args.excel_files:
        excel_path = Path(excel_path_str)
        if not excel_path.exists():
            print(f"ファイルが見つかりません: {excel_path}")
            continue
        if excel_path.suffix.lower() not in {".xlsx", ".xls", ".xlsm"}:
            print(f"Excelファイルではありません（スキップ）: {excel_path}")
            continue

        print(f"処理中: {excel_path}")
        try:
            recipes = parse_excel_file(str(excel_path))
        except Exception as e:
            print(f"  エラー: {e}")
            import traceback
            traceback.print_exc()
            continue

        added = 0
        for recipe in recipes:
            if recipe["title"] not in all_titles:
                all_titles.add(recipe["title"])
                all_recipes.append(recipe)
                added += 1
        print(f"  → {len(recipes)}件抽出, {added}件追加（重複除外済）")

    if not all_recipes:
        print("レシピが1件も抽出できませんでした。")
        print("Excel のフォーマットが対応形式と異なる可能性があります。")
        sys.exit(1)

    # IDを通し番号に再採番
    for i, recipe in enumerate(all_recipes, start=1):
        recipe["id"] = i

    # 出力パスの決定
    if args.output:
        output_path = Path(args.output)
    elif len(args.excel_files) == 1:
        output_path = Path(args.excel_files[0]).with_suffix(".json")
    else:
        output_path = Path("recipes_from_excel.json")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(all_recipes, f, ensure_ascii=False, indent=4)

    print(f"\n完了: {len(all_recipes)}件のレシピを出力しました → {output_path}")
    print()
    print("【出力内容に関する注意】")
    print("  - 栄養素(エネルギー・たんぱく質・脂質)はその日の合計を料理数で按分した概算値です")
    print("  - ナトリウムはExcelに記載がないため 0 になっています")
    print("  - 食材の詳細栄養素(energy/protein/lipid/sodium)はExcelにないため 0 になっています")
    print("  - steps(調理工程数)はデフォルト値 5 を設定しています")
    print("  - 同名レシピが複数月に登場する場合、初出のデータが使われます")
    print("  - 出力JSONを確認し、必要に応じて修正してください")


if __name__ == "__main__":
    main()
